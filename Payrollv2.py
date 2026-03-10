# payroll_export.py - Payroll Export Spreadsheet Generator with On-Call and Training Support

import os
import requests
import datetime
import pytz
import pandas as pd
import streamlit as st
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    cache_data = st.cache_data
except AttributeError:
    cache_data = st.cache

# --- Constants ---
UK_MINIMUM_WAGE = 12.21  # Rate 2 for overtime
ON_CALL_ROLE_NAME = "On-Call"
ON_CALL_FLAT_RATE = 15.00  # £15 per shift
TRAINING_ROLE_NAME = "Training"  # Role name in RotaCloud for training shifts


# --- Helper Functions for Date Management (11th to 10th Monthly Period) ---
def get_monthly_payroll_period(year, month):
    """
    Returns the payroll period from 11th of given month to 10th of next month.
    Returns (start_date, end_date) as datetime.date objects.
    """
    try:
        start_date = datetime.date(year, month, 11)
    except ValueError:
        st.error(f"Error: Invalid date for start of period: Year {year}, Month {month}, Day 11.")
        return None, None
    
    if month == 12:
        next_month = 1
        next_year = year + 1
    else:
        next_month = month + 1
        next_year = year
    
    try:
        end_date = datetime.date(next_year, next_month, 10)
    except ValueError:
        st.error(f"Error: Invalid date for end of period: Year {next_year}, Month {next_month}, Day 10.")
        return None, None
        
    return start_date, end_date

def get_default_payroll_period():
    """
    Returns the most recently completed payroll period.
    If today is on or before the 10th, use the period that ended on 10th of current month.
    If today is after the 10th, use the period that ended on 10th of current month.
    """
    today = datetime.date.today()
    
    if today.day <= 10:
        if today.month == 1:
            target_month = 12
            target_year = today.year - 1
        else:
            target_month = today.month - 1
            target_year = today.year
    else:
        if today.month == 1:
            target_month = 12
            target_year = today.year - 1
        else:
            target_month = today.month - 1
            target_year = today.year
    
    return get_monthly_payroll_period(target_year, target_month)

def format_payroll_period_name(start_date, end_date):
    """Format payroll period name for display."""
    if start_date and end_date:
        return f"{start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}"
    return "Invalid Period"

def calculate_period_days(start_date, end_date):
    """Calculate the number of days in the period."""
    return (end_date - start_date).days + 1

# Initialize session state for period management
if 'payroll_period_year' not in st.session_state:
    current_start, current_end = get_default_payroll_period()
    if current_start:
        st.session_state.payroll_period_year = current_start.year
        st.session_state.payroll_period_month = current_start.month
    else:
        st.session_state.payroll_period_year = datetime.date.today().year
        st.session_state.payroll_period_month = datetime.date.today().month

# --- Sidebar: Configuration Inputs ---
st.sidebar.header("Payroll Export Configuration")

# Debug mode toggle
DEBUG_MODE = st.sidebar.checkbox("🐛 Enable Debug Mode", value=False, help="Show detailed debugging information")

st.sidebar.subheader("Payroll Period (11th - 10th)")

# Calculate current period
current_start, current_end = get_monthly_payroll_period(
    st.session_state.payroll_period_year, 
    st.session_state.payroll_period_month
)

if current_start and current_end:
    period_name = format_payroll_period_name(current_start, current_end)
    period_days = calculate_period_days(current_start, current_end)
    st.sidebar.write(f"**Current Period:** {period_name}")
    st.sidebar.write(f"**Days in Period:** {period_days}")

# Navigation buttons
col1, col2, col3 = st.sidebar.columns(3)

with col1:
    if st.button("◀ Previous", help="Go to previous payroll period"):
        if st.session_state.payroll_period_month == 1:
            st.session_state.payroll_period_month = 12
            st.session_state.payroll_period_year -= 1
        else:
            st.session_state.payroll_period_month -= 1
        st.rerun()

with col2:
    if st.button("Default", help="Go to most recently completed period"):
        default_start, default_end = get_default_payroll_period()
        if default_start:
            st.session_state.payroll_period_year = default_start.year
            st.session_state.payroll_period_month = default_start.month
        st.rerun()

with col3:
    if st.button("Next ▶", help="Go to next payroll period"):
        if st.session_state.payroll_period_month == 12:
            st.session_state.payroll_period_month = 1
            st.session_state.payroll_period_year += 1
        else:
            st.session_state.payroll_period_month += 1
        st.rerun()

# Get the dates for the selected period
start_date, end_date = get_monthly_payroll_period(
    st.session_state.payroll_period_year, 
    st.session_state.payroll_period_month
)

if start_date is None or end_date is None:
    st.stop()

# API Key input
api_key_input = st.sidebar.text_input(
    "Rotacloud API Key", 
    type="password",
    help="Enter your Rotacloud API key"
)

# Ignored user IDs
ignored_input = st.sidebar.text_input(
    "Ignored User IDs (comma-separated)", 
    value=""
)
ignored_user_ids = set()
for part in ignored_input.split(","):
    part = part.strip()
    if part.isdigit():
        ignored_user_ids.add(int(part))

# Overtime rate override
overtime_rate = st.sidebar.number_input(
    "Overtime Rate (Rate 2) £/hr",
    value=UK_MINIMUM_WAGE,
    step=0.01,
    format="%.2f",
    help="Rate applied to hours exceeding contracted hours (default: UK minimum wage)"
)

# Generate button
generate_report_button = st.sidebar.button("Generate Payroll Export", type="primary")

# --- Main Title ---
st.title("📊 Payroll Export Spreadsheet Generator")

if current_start and current_end:
    st.info(f"📅 **Payroll Period**: {format_payroll_period_name(current_start, current_end)} ({calculate_period_days(current_start, current_end)} days)")

# --- API Setup ---
API_KEY = None

if api_key_input:
    API_KEY = api_key_input.strip()
elif os.environ.get("ROTACLOUD_API_KEY"):
    API_KEY = os.environ.get("ROTACLOUD_API_KEY")

if not API_KEY:
    st.error("⚠️ **API Key Required**")
    st.write("Please provide your Rotacloud API key in the sidebar.")
    st.stop()

HEADERS = {"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"}

# --- Base URLs ---
USERS_BASE_URL = "https://api.rotacloud.com/v1/users"
SHIFTS_BASE_URL = "https://api.rotacloud.com/v1/shifts"
LEAVE_BASE_URL = "https://api.rotacloud.com/v1/leave"
ROLES_BASE_URL = "https://api.rotacloud.com/v1/roles"
ATTENDANCE_BASE_URL = "https://api.rotacloud.com/v1/attendance"

# --- Helper Functions ---
def date_to_unix_timestamp(date_obj, hour=0, minute=0, second=0, timezone_str='Europe/London'):
    if not date_obj:
        return None
    try:
        local_tz = pytz.timezone(timezone_str)
        dt_naive = datetime.datetime.combine(date_obj, datetime.time(hour, minute, second))
        local_dt = local_tz.localize(dt_naive, is_dst=None)
        utc_dt = local_dt.astimezone(pytz.utc)
        return int(utc_dt.timestamp())
    except Exception as e:
        st.warning(f"Error converting date '{date_obj}' to timestamp: {e}")
        return None

def date_str_to_date(date_str):
    if not date_str:
        return None
    try:
        return datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return None

@cache_data(ttl=300)
def get_rotacloud_users():
    try:
        resp = requests.get(USERS_BASE_URL, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.HTTPError as e:
        st.error(f"HTTP Error fetching users: {e}")
        return None
    except requests.exceptions.RequestException as e:
        st.error(f"Connection Error: {e}")
        return None

@cache_data(ttl=300)
def get_rotacloud_shifts(start_ts, end_ts, user_id):
    params = {
        "start": start_ts,
        "end": end_ts,
        "published": "true"
    }
    params["users[]"] = [user_id]
    try:
        resp = requests.get(SHIFTS_BASE_URL, headers=HEADERS, params=params, timeout=30)
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.RequestException as e:
        return None

@cache_data(ttl=300)
def get_rotacloud_leave(start_str, end_str, user_id):
    params = {
        "start": start_str,
        "end": end_str,
        "include_deleted": "false",
        "include_denied": "false",
        "include_requested": "false",
        "include_expired": "true"
    }
    params["users[]"] = [user_id]
    try:
        resp = requests.get(LEAVE_BASE_URL, headers=HEADERS, params=params, timeout=30)
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.RequestException:
        return None

@cache_data(ttl=300)
def get_attendance_data(start_ts, end_ts, user_id):
    """Fetch attendance records for a user in the period (for On-Call hour calculations)."""
    params = {
        "start": start_ts,
        "end": end_ts
    }
    params["users[]"] = [user_id]
    try:
        resp = requests.get(ATTENDANCE_BASE_URL, headers=HEADERS, params=params, timeout=30)
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.RequestException as e:
        st.warning(f"Warning: Could not fetch attendance data for user {user_id}: {e}")
        return None

@cache_data(ttl=600)
def get_role_name(role_id):
    if role_id is None:
        return "Unknown Role"
    try:
        resp = requests.get(f"{ROLES_BASE_URL}/{role_id}", headers=HEADERS, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        return data.get("name", f"Role {role_id}")
    except:
        return f"Role {role_id}"

def get_user_pay_details(user_data):
    pay_info = {
        "pay_type": user_data.get("salary_type"),
        "hourly_rate": float(user_data.get("salary")) if user_data.get("salary_type") == "hourly" and user_data.get("salary") is not None else 0.0,
        "annual_salary": float(user_data.get("salary")) if user_data.get("salary_type") == "annual" and user_data.get("salary") is not None else 0.0,
        "role_rates": {},
        "standard_weekly_hours": float(user_data.get("weekly_hours")) if user_data.get("weekly_hours") is not None else 0.0
    }
    
    api_role_rates = user_data.get("role_rates")
    if api_role_rates:
        for role_id_str, rate_data in api_role_rates.items():
            try:
                role_id_int = int(role_id_str)
                per_hour_rate = rate_data.get("per_hour")
                if per_hour_rate is not None:
                    pay_info["role_rates"][role_id_int] = float(per_hour_rate)
            except (ValueError, AttributeError):
                pass
    
    return pay_info

def calculate_shift_hours_by_role(shifts_json, user_pay_details, default_hourly_rate, all_custom_roles, role_id_to_name):
    """
    Calculate hours worked, broken down by role rate.
    EXCLUDES On-Call and Training shifts from the main hour calculations (they are handled separately).
    Returns: total_hours, hours_at_base_rate, dict of {role_id: (hours, rate, role_name)},
             on_call_shifts_list, training_hours (float)
    """
    total_seconds = 0
    base_rate_seconds = 0
    role_hours = {}  # {role_id: {'seconds': 0, 'rate': rate, 'name': name}}
    on_call_shifts = []  # List of On-Call shift records for attendance lookup
    training_seconds = 0  # Training hours calculated from scheduled shift times

    if not shifts_json:
        return 0.0, 0.0, {}, on_call_shifts, 0.0

    role_rates = user_pay_details.get("role_rates", {})

    if DEBUG_MODE:
        st.write(f"🔍 **DEBUG: Processing {len(shifts_json)} shifts**")

    for shift in shifts_json:
        st_unix = shift.get("start_time")
        en_unix = shift.get("end_time")
        minutes_break = shift.get("minutes_break", 0) or 0
        role_id = shift.get("role")
        shift_id = shift.get("id")

        # Get role name for this shift
        if role_id not in role_id_to_name:
            role_id_to_name[role_id] = get_role_name(role_id)
        role_name = role_id_to_name.get(role_id, "Unknown")

        if st_unix is None or en_unix is None:
            continue

        # --- On-Call: exclude from normal hours, collect for attendance lookup ---
        if role_name == ON_CALL_ROLE_NAME:
            if DEBUG_MODE:
                st.write(f"   ✅ Found On-Call shift: ID={shift_id}, Role={role_name}")
            on_call_shifts.append({
                "shift_id": shift_id,
                "start_time": st_unix,
                "end_time": en_unix,
                "minutes_break": minutes_break
            })
            continue

        # --- Training: exclude from normal hours, accumulate from scheduled times ---
        if role_name == TRAINING_ROLE_NAME:
            duration = en_unix - st_unix
            net_seconds = duration - minutes_break * 60
            training_seconds += net_seconds
            if DEBUG_MODE:
                st.write(f"   📚 Found Training shift: ID={shift_id}, hours={net_seconds/3600:.2f}")
            continue

        # --- Regular shift ---
        if DEBUG_MODE and role_name:
            st.write(f"   ℹ️ Regular shift: ID={shift_id}, Role={role_name}")

        duration = en_unix - st_unix
        net_seconds = duration - minutes_break * 60
        total_seconds += net_seconds
        
        # Check if this role has a custom rate
        if role_id and role_id in role_rates:
            custom_rate = role_rates[role_id]
            if role_id not in role_hours:
                role_name_fetched = get_role_name(role_id)
                role_hours[role_id] = {'seconds': 0, 'rate': custom_rate, 'name': role_name_fetched}
                all_custom_roles[role_id] = role_name_fetched
            role_hours[role_id]['seconds'] += net_seconds
        else:
            base_rate_seconds += net_seconds
    
    # Convert to hours
    total_hours = round(total_seconds / 3600, 2)
    base_rate_hours = round(base_rate_seconds / 3600, 2)
    training_hours = round(training_seconds / 3600, 2)
    
    role_hours_converted = {}
    for role_id, data in role_hours.items():
        role_hours_converted[role_id] = {
            'hours': round(data['seconds'] / 3600, 2),
            'rate': data['rate'],
            'name': data['name']
        }
    
    return total_hours, base_rate_hours, role_hours_converted, on_call_shifts, training_hours

def unix_to_datetime(unix_ts, timezone_str='Europe/London'):
    """Convert Unix timestamp to timezone-aware datetime."""
    if not unix_ts:
        return None
    try:
        utc_dt = datetime.datetime.fromtimestamp(unix_ts, tz=pytz.utc)
        local_tz = pytz.timezone(timezone_str)
        return utc_dt.astimezone(local_tz)
    except Exception:
        return None

def calculate_on_call_hours(on_call_shifts, attendance_data):
    """
    Calculate On-Call hours from attendance data using in_time and out_time.
    Returns: total_on_call_hours, number_of_on_call_shifts (all assigned, not just attended)
    """
    if DEBUG_MODE:
        st.write(f"🔍 **DEBUG: calculate_on_call_hours called**")
        st.write(f"   - On-Call shifts provided: {len(on_call_shifts) if on_call_shifts else 0}")
        st.write(f"   - Attendance records provided: {len(attendance_data) if attendance_data else 0}")

    if not on_call_shifts:
        if DEBUG_MODE:
            st.write("   ⚠️ No On-Call shifts to process")
        return 0.0, 0

    total_on_call_shifts = len(on_call_shifts)

    if not attendance_data:
        if DEBUG_MODE:
            st.write("   ⚠️ No attendance data to process")
        return 0.0, total_on_call_shifts

    total_on_call_hours = 0.0

    shift_lookup = {}
    for shift in on_call_shifts:
        shift_id = shift.get("shift_id")
        if not shift_id:
            continue
        shift_lookup[shift_id] = {
            "start_time": shift.get("start_time"),
            "end_time": shift.get("end_time"),
            "minutes_break": shift.get("minutes_break", 0) or 0
        }

    if DEBUG_MODE:
        st.write(f"   📋 Built lookup for {len(shift_lookup)} On-Call shifts: {list(shift_lookup.keys())}")

    for attendance in attendance_data:
        if attendance.get("deleted"):
            continue

        shift_id = attendance.get("shift")
        in_time = attendance.get("in_time")
        out_time = attendance.get("out_time")

        if DEBUG_MODE:
            st.write(f"   🔎 Checking attendance: shift_id={shift_id}, in_time={in_time}, out_time={out_time}")

        if shift_id not in shift_lookup:
            if DEBUG_MODE:
                st.write(f"      ⏭️ Shift {shift_id} not in On-Call lookup")
            continue

        if DEBUG_MODE:
            st.write(f"      ✅ MATCHED On-Call shift {shift_id}!")

        if not in_time or not out_time:
            if DEBUG_MODE:
                st.write(f"      ⚠️ Missing in_time or out_time")
            continue

        try:
            in_time_dt = unix_to_datetime(in_time)
            out_time_dt = unix_to_datetime(out_time)

            if not in_time_dt or not out_time_dt:
                st.warning(f"Could not convert timestamps for On-Call shift {shift_id}")
                continue

            hours_worked = (out_time - in_time) / 3600.0

            if DEBUG_MODE:
                st.write(f"      📊 Hours calculated: {hours_worked:.2f}")
                st.write(f"         Clock in:  {in_time_dt.strftime('%Y-%m-%d %H:%M:%S')}")
                st.write(f"         Clock out: {out_time_dt.strftime('%Y-%m-%d %H:%M:%S')}")

            if hours_worked <= 0:
                st.warning(f"Invalid On-Call hours for shift {shift_id}: {hours_worked:.2f} hours")
                continue

            if hours_worked > 12:
                st.warning(f"Suspiciously long On-Call shift {shift_id}: {hours_worked:.2f} hours")

            total_on_call_hours += hours_worked

            if DEBUG_MODE:
                st.write(f"      ✅ Added to totals. Running total: {total_on_call_hours:.2f} hours")

        except (TypeError, ValueError) as e:
            st.warning(f"Could not calculate On-Call hours for shift {shift_id}: {e}")

    if DEBUG_MODE:
        st.write(f"   🎯 **FINAL RESULT: {total_on_call_hours:.2f} hours from {total_on_call_shifts} assigned shifts**")

    return round(total_on_call_hours, 2), total_on_call_shifts

def calculate_leave_hours(leave_json, report_start_date, report_end_date):
    """Calculate total approved holiday and sickness within the period."""
    holiday_days = 0.0
    holiday_hours = 0.0
    sickness_days = 0.0
    
    if not leave_json:
        return 0.0, 0.0, 0.0
    
    for record in leave_json:
        if record.get("status") != "approved":
            continue
        
        leave_type = record.get("type")
        
        for date_entry in record.get("dates", []):
            dstr = date_entry.get("date")
            ddate = date_str_to_date(dstr)
            if not ddate:
                continue
            if report_start_date <= ddate <= report_end_date:
                days_val = date_entry.get("days", 0) or 0
                hours_val = date_entry.get("hours", 0) or 0
                
                if leave_type == 1:  # Holiday
                    holiday_days += days_val
                    holiday_hours += hours_val
                elif leave_type == 3:  # Sickness
                    sickness_days += days_val
    
    return round(holiday_days, 2), round(holiday_hours, 2), round(sickness_days, 2)

def calculate_fixed_hours(weekly_hours, period_days):
    """
    Calculate fixed contracted hours for the period.
    Formula: (Full Weeks * Weekly Hours) + ((Remainder Days / 7) * Weekly Hours)
    """
    if weekly_hours <= 0 or period_days <= 0:
        return 0.0
    
    full_weeks = period_days // 7
    remainder_days = period_days % 7
    full_weeks_total = full_weeks * weekly_hours
    proportional_remainder = (remainder_days / 7) * weekly_hours
    fixed_hours = full_weeks_total + proportional_remainder
    
    return round(fixed_hours, 2)


def create_payroll_excel(payroll_data, start_date, end_date, overtime_rate, all_custom_roles):
    """Create Excel workbook with payroll data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Export"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    salaried_fill = PatternFill("solid", fgColor="E2EFDA")        # Light green
    non_standard_rate_fill = PatternFill("solid", fgColor="FCE4D6")  # Light orange
    on_call_fill = PatternFill("solid", fgColor="FFF2CC")          # Light yellow
    training_fill = PatternFill("solid", fgColor="E8D5F5")         # Light purple for training
    currency_format = '£#,##0.00'
    number_format = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Build dynamic headers
    base_headers = [
        "Employee Name",
        "Pay Type",
        "Weekly Hrs",
        "Total Hrs",
        "Fixed Hrs",
        "Hours",
        "Rate 1 (£)",
    ]
    
    # Custom role columns
    custom_role_headers = []
    sorted_roles = sorted(all_custom_roles.items(), key=lambda x: x[1])
    for role_id, role_name in sorted_roles:
        custom_role_headers.append(f"{role_name} Hrs")
        custom_role_headers.append(f"{role_name} Rate (£)")
    
    # On-Call columns
    on_call_headers = [
        "On-Call Hrs",
        "On-Call Shifts",
        "On-Call Flat Rate (£)"
    ]

    # Training columns — hours from scheduled shift times, paid at Rate 1
    training_headers = [
        "Training Hrs",
        "Training Pay (£)"  # Calculated: Training Hrs × Rate 1 (bold, formula)
    ]
    
    end_headers = [
        "Overtime Hrs",
        "Rate 2 (£)",
        "Holiday (Days)",
        "Holiday (Hrs)",
        "Sickness (Days)",
        "TOTAL PAY (£)"
    ]
    
    headers = base_headers + custom_role_headers + on_call_headers + training_headers + end_headers
    
    # Row 1: Period header
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    period_cell = ws.cell(row=1, column=1)
    period_cell.value = f"Payroll Period: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
    period_cell.font = Font(bold=True, size=14)
    period_cell.alignment = Alignment(horizontal='center')
    
    # Row 2: Column headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 10
    for col_idx in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    
    # --- Column index assignments ---
    weekly_hrs_col = 3   # C
    total_hrs_col = 4    # D
    fixed_hrs_col = 5    # E
    hours_col = 6        # F
    rate1_col = 7        # G

    custom_role_start_col = 8
    num_custom_role_cols = len(custom_role_headers)

    on_call_hrs_col       = custom_role_start_col + num_custom_role_cols
    on_call_shifts_col    = on_call_hrs_col + 1
    on_call_flat_rate_col = on_call_shifts_col + 1

    training_hrs_col  = on_call_flat_rate_col + 1
    training_pay_col  = training_hrs_col + 1   # Calculated: Training Hrs × Rate 1

    overtime_hrs_col  = training_pay_col + 1
    rate2_col         = overtime_hrs_col + 1
    holiday_days_col  = rate2_col + 1
    holiday_hrs_col   = holiday_days_col + 1
    sickness_days_col = holiday_hrs_col + 1
    total_pay_col     = sickness_days_col + 1

    STANDARD_RATE = 12.21
    data_start_row = 3
    
    for row_offset, data in enumerate(payroll_data):
        row_idx = data_start_row + row_offset
        is_salaried = data['pay_type'] == 'annual'
        is_non_standard_rate = not is_salaried and abs(data['rate_1'] - STANDARD_RATE) > 0.01
        has_on_call = data['on_call_hours'] > 0
        has_training = data['training_hours'] > 0

        # Row highlight priority: on-call > training > salaried > non-standard rate
        if has_on_call:
            row_fill = on_call_fill
        elif has_training:
            row_fill = training_fill
        elif is_salaried:
            row_fill = salaried_fill
        elif is_non_standard_rate:
            row_fill = non_standard_rate_fill
        else:
            row_fill = None

        def styled_cell(row, col, value, fmt=None, bold=False):
            c = ws.cell(row=row, column=col, value=value)
            c.border = thin_border
            if row_fill:
                c.fill = row_fill
            if fmt:
                c.number_format = fmt
            if bold:
                c.font = Font(bold=True)
            return c

        # Basic columns
        styled_cell(row_idx, 1, data['employee_name'])
        styled_cell(row_idx, 2, "Salaried" if is_salaried else "Hourly")
        styled_cell(row_idx, weekly_hrs_col, data['weekly_hours'], number_format)
        styled_cell(row_idx, total_hrs_col, data['total_hours_display'], number_format)
        styled_cell(row_idx, fixed_hrs_col, data['fixed_hours'], number_format, bold=True)

        # Hours = MIN(Total - On-Call, Fixed - On-Call)
        hours_formula = (
            f"=MIN({get_column_letter(total_hrs_col)}{row_idx}-{get_column_letter(on_call_hrs_col)}{row_idx}"
            f"-{get_column_letter(training_hrs_col)}{row_idx},"
            f"{get_column_letter(fixed_hrs_col)}{row_idx}-{get_column_letter(on_call_hrs_col)}{row_idx}"
            f"-{get_column_letter(training_hrs_col)}{row_idx})"
        )
        c = ws.cell(row=row_idx, column=hours_col, value=hours_formula)
        c.border = thin_border
        c.number_format = number_format
        c.font = Font(bold=True)
        if row_fill:
            c.fill = row_fill

        styled_cell(row_idx, rate1_col, data['rate_1'], currency_format)

        # Custom role columns
        col_offset = 0
        for role_id, role_name in sorted_roles:
            hrs_col = custom_role_start_col + col_offset
            rate_col = custom_role_start_col + col_offset + 1
            role_data = data['custom_role_hours'].get(role_id, {'hours': 0, 'rate': 0})
            styled_cell(row_idx, hrs_col, role_data['hours'], number_format)
            styled_cell(row_idx, rate_col, role_data['rate'], currency_format)
            col_offset += 2

        # On-Call columns
        styled_cell(row_idx, on_call_hrs_col, data['on_call_hours'], number_format, bold=True)
        styled_cell(row_idx, on_call_shifts_col, data['on_call_shift_count'], number_format)
        styled_cell(row_idx, on_call_flat_rate_col, ON_CALL_FLAT_RATE, currency_format)

        # Training columns
        styled_cell(row_idx, training_hrs_col, data['training_hours'], number_format, bold=True)
        # Training Pay = Training Hrs × Rate 1 (calculated formula, bold)
        training_pay_formula = (
            f"={get_column_letter(training_hrs_col)}{row_idx}*{get_column_letter(rate1_col)}{row_idx}"
        )
        c = ws.cell(row=row_idx, column=training_pay_col, value=training_pay_formula)
        c.border = thin_border
        c.number_format = currency_format
        c.font = Font(bold=True)
        if row_fill:
            c.fill = row_fill

        # Overtime = MAX(0, Total - Fixed)
        overtime_formula = (
            f"=MAX(0,{get_column_letter(total_hrs_col)}{row_idx}"
            f"-{get_column_letter(fixed_hrs_col)}{row_idx})"
        )
        c = ws.cell(row=row_idx, column=overtime_hrs_col, value=overtime_formula)
        c.border = thin_border
        c.number_format = number_format
        c.font = Font(bold=True)
        if row_fill:
            c.fill = row_fill

        styled_cell(row_idx, rate2_col, overtime_rate, currency_format)
        styled_cell(row_idx, holiday_days_col, data['holiday_days'], number_format)
        styled_cell(row_idx, holiday_hrs_col, data['holiday_hours'], number_format)
        styled_cell(row_idx, sickness_days_col, data['sickness_days'], number_format)

        # TOTAL PAY
        if is_salaried:
            monthly_salary = data['annual_salary'] / 12
            c = ws.cell(row=row_idx, column=total_pay_col, value=monthly_salary)
        else:
            hours_letter          = get_column_letter(hours_col)
            rate1_letter          = get_column_letter(rate1_col)
            on_call_hrs_letter    = get_column_letter(on_call_hrs_col)
            on_call_shifts_letter = get_column_letter(on_call_shifts_col)
            on_call_flat_letter   = get_column_letter(on_call_flat_rate_col)
            training_hrs_letter   = get_column_letter(training_hrs_col)
            overtime_hrs_letter   = get_column_letter(overtime_hrs_col)
            rate2_letter          = get_column_letter(rate2_col)
            holiday_hrs_letter    = get_column_letter(holiday_hrs_col)

            custom_role_pay_parts = []
            col_offset = 0
            for role_id, role_name in sorted_roles:
                h = get_column_letter(custom_role_start_col + col_offset)
                r = get_column_letter(custom_role_start_col + col_offset + 1)
                custom_role_pay_parts.append(f"({h}{row_idx}*{r}{row_idx})")
                col_offset += 2

            custom_role_pay_formula = "+".join(custom_role_pay_parts) if custom_role_pay_parts else ""

            # Total = (Hours × Rate1) + Custom Roles + (On-Call Hrs × Rate1) + (On-Call Shifts × Flat)
            #       + (Training Hrs × Rate1) + (Overtime × Rate2) + (Holiday Hrs × Rate1)
            total_pay_formula = f"=({hours_letter}{row_idx}*{rate1_letter}{row_idx})"
            if custom_role_pay_formula:
                total_pay_formula += f"+{custom_role_pay_formula}"
            total_pay_formula += (
                f"+({on_call_hrs_letter}{row_idx}*{rate1_letter}{row_idx})"
                f"+({on_call_shifts_letter}{row_idx}*{on_call_flat_letter}{row_idx})"
                f"+({training_hrs_letter}{row_idx}*{rate1_letter}{row_idx})"
                f"+({overtime_hrs_letter}{row_idx}*{rate2_letter}{row_idx})"
                f"+({holiday_hrs_letter}{row_idx}*{rate1_letter}{row_idx})"
            )
            c = ws.cell(row=row_idx, column=total_pay_col, value=total_pay_formula)

        c.border = thin_border
        c.number_format = currency_format
        c.font = Font(bold=True)
        if row_fill:
            c.fill = row_fill

    # Totals row
    last_data_row = data_start_row + len(payroll_data) - 1
    total_row = last_data_row + 1

    cell = ws.cell(row=total_row, column=1, value="TOTALS")
    cell.font = Font(bold=True)
    cell.border = thin_border

    sum_columns = [weekly_hrs_col, total_hrs_col, fixed_hrs_col, hours_col]
    col_offset = 0
    for role_id, role_name in sorted_roles:
        sum_columns.append(custom_role_start_col + col_offset)
        col_offset += 2
    sum_columns.extend([
        on_call_hrs_col, on_call_shifts_col,
        training_hrs_col,
        overtime_hrs_col,
        holiday_days_col, holiday_hrs_col, sickness_days_col,
        total_pay_col
    ])

    for col in sum_columns:
        col_letter = get_column_letter(col)
        cell = ws.cell(row=total_row, column=col, value=f"=SUM({col_letter}{data_start_row}:{col_letter}{last_data_row})")
        cell.border = thin_border
        cell.font = Font(bold=True)
        if col == total_pay_col:
            cell.number_format = currency_format
            cell.fill = PatternFill("solid", fgColor="FFFF00")
        elif col in [rate1_col, rate2_col, on_call_flat_rate_col]:
            cell.number_format = currency_format
        else:
            cell.number_format = number_format

    ws.row_dimensions[2].height = 30

    return wb


# --- Main Report Generation ---
if generate_report_button:
    st.info("Generating payroll export... this may take some time.")
    
    start_str = start_date.strftime("%Y-%m-%d")
    end_str = end_date.strftime("%Y-%m-%d")
    start_ts = date_to_unix_timestamp(start_date)
    end_ts = date_to_unix_timestamp(end_date, hour=23, minute=59, second=59)
    period_days = calculate_period_days(start_date, end_date)
    
    if start_ts is None or end_ts is None:
        st.error("Date conversion error.")
        st.stop()
    
    with st.spinner("Testing API connection..."):
        try:
            test_resp = requests.get(USERS_BASE_URL, headers=HEADERS, params={"limit": 1}, timeout=10)
            test_resp.raise_for_status()
            st.success("✅ API Connection Successful")
        except requests.exceptions.RequestException as e:
            st.error(f"❌ API Connection Failed: {e}")
            st.stop()
    
    all_users = get_rotacloud_users()
    if all_users is None:
        st.error("Failed to fetch user data.")
        st.stop()
    
    role_id_to_name = {}
    payroll_data = []
    all_custom_roles = {}
    progress_bar = st.progress(0, text="Processing users...")
    
    users_to_process = [u for u in all_users if u.get("id") not in ignored_user_ids]
    total_users = len(users_to_process)
    
    for idx, user in enumerate(users_to_process):
        user_id = user.get("id")
        first_name = user.get("first_name", "")
        last_name = user.get("last_name", "")
        employee_name = f"{first_name} {last_name}".strip()
        
        progress_bar.progress((idx + 1) / total_users, text=f"Processing {employee_name}...")
        
        pay_details = get_user_pay_details(user)
        pay_type = pay_details["pay_type"]
        
        if pay_type == "hourly":
            rate_1 = pay_details["hourly_rate"]
        elif pay_type == "annual" and pay_details["standard_weekly_hours"] > 0:
            rate_1 = (pay_details["annual_salary"] / 52) / pay_details["standard_weekly_hours"]
        else:
            rate_1 = 0.0
        
        weekly_hours = pay_details["standard_weekly_hours"]
        fixed_hours = calculate_fixed_hours(weekly_hours, period_days)
        
        if DEBUG_MODE:
            st.write(f"\n---\n### 👤 Processing: **{employee_name}** (ID: {user_id})")

        shifts_json = get_rotacloud_shifts(start_ts, end_ts, user_id)

        if DEBUG_MODE:
            st.write(f"📅 **Fetched {len(shifts_json) if shifts_json else 0} shifts for {employee_name}**")

        total_hours, base_rate_hours, custom_role_hours, on_call_shifts, training_hours = calculate_shift_hours_by_role(
            shifts_json, pay_details, rate_1, all_custom_roles, role_id_to_name
        )

        if DEBUG_MODE:
            st.write(f"📊 **Shift breakdown for {employee_name}:**")
            st.write(f"   - Total hours (excl. On-Call & Training): {total_hours}")
            st.write(f"   - Training hours: {training_hours}")
            st.write(f"   - On-Call shifts found: {len(on_call_shifts)}")

        for role_id, role_data in custom_role_hours.items():
            if role_id not in all_custom_roles:
                all_custom_roles[role_id] = role_data['name']

        on_call_hours = 0.0
        on_call_shift_count = 0
        if on_call_shifts:
            if DEBUG_MODE:
                st.write(f"🚨 **Fetching attendance data for On-Call shifts...**")
            attendance_data = get_attendance_data(start_ts, end_ts, user_id)
            if DEBUG_MODE and attendance_data:
                st.write(f"📋 **Fetched {len(attendance_data)} attendance records**")
                st.json({"sample_attendance": attendance_data[0] if attendance_data else "No data"})
            on_call_hours, on_call_shift_count = calculate_on_call_hours(on_call_shifts, attendance_data)
            if DEBUG_MODE:
                st.write(f"✅ **On-Call Result: {on_call_hours} hours from {on_call_shift_count} shifts**")
        elif DEBUG_MODE:
            st.write(f"ℹ️ No On-Call shifts for {employee_name}")

        leave_json = get_rotacloud_leave(start_str, end_str, user_id)
        holiday_days, holiday_hours, sickness_days = calculate_leave_hours(leave_json, start_date, end_date)
        
        if total_hours > 0 or holiday_hours > 0 or holiday_days > 0 or sickness_days > 0 or on_call_hours > 0 or training_hours > 0:
            payroll_data.append({
                "employee_name": employee_name,
                "pay_type": pay_type,
                "annual_salary": pay_details["annual_salary"],
                "hourly_rate": pay_details["hourly_rate"],
                "weekly_hours": pay_details["standard_weekly_hours"],
                "total_hours": total_hours,
                "total_hours_display": total_hours + on_call_hours + training_hours,
                "fixed_hours": fixed_hours,
                "rate_1": round(rate_1, 2),
                "base_rate_hours": base_rate_hours,
                "custom_role_hours": custom_role_hours,
                "on_call_hours": on_call_hours,
                "on_call_shift_count": on_call_shift_count,
                "training_hours": training_hours,
                "holiday_days": holiday_days,
                "holiday_hours": holiday_hours,
                "sickness_days": sickness_days
            })
    
    progress_bar.empty()
    
    if not payroll_data:
        st.warning("No payroll data found for the selected period.")
        st.stop()
    
    salaried_staff = [p for p in payroll_data if p['pay_type'] == 'annual']
    hourly_staff = [p for p in payroll_data if p['pay_type'] != 'annual']
    salaried_staff.sort(key=lambda x: x['employee_name'].lower())
    hourly_staff.sort(key=lambda x: x['employee_name'].lower())
    payroll_data = salaried_staff + hourly_staff
    
    st.subheader("📋 Payroll Preview")
    
    if all_custom_roles:
        st.info(f"🏷️ **Custom Role Rates Found:** {', '.join(all_custom_roles.values())}")
    
    preview_rows = []
    for data in payroll_data:
        hours_worked = min(
            data['total_hours'],
            data['fixed_hours'] - data['on_call_hours'] - data['training_hours']
        )
        overtime_hrs = max(0, data['total_hours_display'] - data['fixed_hours'])

        if data['pay_type'] == 'annual':
            total_pay = data['annual_salary'] / 12
        else:
            custom_role_pay = sum(rd['hours'] * rd['rate'] for rd in data['custom_role_hours'].values())
            base_pay        = hours_worked * data['rate_1']
            on_call_hrs_pay = data['on_call_hours'] * data['rate_1']
            on_call_flat_pay = data['on_call_shift_count'] * ON_CALL_FLAT_RATE
            training_pay    = data['training_hours'] * data['rate_1']
            overtime_pay    = overtime_hrs * overtime_rate
            holiday_pay     = data['holiday_hours'] * data['rate_1']
            total_pay = base_pay + custom_role_pay + on_call_hrs_pay + on_call_flat_pay + training_pay + overtime_pay + holiday_pay
        
        row = {
            'Employee': data['employee_name'],
            'Type': 'Salaried' if data['pay_type'] == 'annual' else 'Hourly',
            'Weekly Hrs': data['weekly_hours'],
            'Total Hrs': data['total_hours_display'],
            'Fixed Hrs': data['fixed_hours'],
            'Hours': hours_worked,
            'Rate 1 (£)': data['rate_1'],
        }
        
        for role_id, role_name in sorted(all_custom_roles.items(), key=lambda x: x[1]):
            role_data = data['custom_role_hours'].get(role_id, {'hours': 0, 'rate': 0})
            row[f'{role_name} Hrs'] = role_data['hours']
        
        row['On-Call Hrs']   = data['on_call_hours']
        row['On-Call Shifts'] = data['on_call_shift_count']
        row['Training Hrs']  = data['training_hours']
        row['Training Pay (£)'] = round(data['training_hours'] * data['rate_1'], 2)
        row['Overtime Hrs']  = overtime_hrs
        row['Holiday (Days)'] = data['holiday_days']
        row['Holiday (Hrs)'] = data['holiday_hours']
        row['Sickness (Days)'] = data['sickness_days']
        row['Total Pay (£)'] = round(total_pay, 2)
        
        preview_rows.append(row)
    
    preview_df = pd.DataFrame(preview_rows)
    st.dataframe(preview_df, use_container_width=True)
    
    total_pay_sum          = sum(r['Total Pay (£)'] for r in preview_rows)
    total_hours_sum        = sum(r['Total Hrs'] for r in preview_rows)
    total_on_call_hrs_sum  = sum(r['On-Call Hrs'] for r in preview_rows)
    total_on_call_shft_sum = sum(r['On-Call Shifts'] for r in preview_rows)
    total_training_hrs_sum = sum(r['Training Hrs'] for r in preview_rows)
    total_overtime_sum     = sum(r['Overtime Hrs'] for r in preview_rows)
    
    col1, col2, col3, col4, col5, col6 = st.columns(6)
    with col1:
        st.metric("Total Employees", len(payroll_data))
    with col2:
        st.metric("Total Hours", f"{total_hours_sum:.1f}")
    with col3:
        st.metric("On-Call Hours", f"{total_on_call_hrs_sum:.1f}")
    with col4:
        st.metric("On-Call Shifts", int(total_on_call_shft_sum))
    with col5:
        st.metric("Training Hours", f"{total_training_hrs_sum:.1f}")
    with col6:
        st.metric("Total Payroll", f"£{total_pay_sum:,.2f}")
    
    wb = create_payroll_excel(payroll_data, start_date, end_date, overtime_rate, all_custom_roles)
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    filename = f"payroll_export_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    
    st.download_button(
        label="📥 Download Payroll Export (Excel)",
        data=excel_buffer.getvalue(),
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )
    
    st.success("✅ Payroll export generated successfully!")

else:
    st.write("Configure the payroll period in the sidebar and click **Generate Payroll Export** to begin.")
    
    st.markdown("""
    ### How it works:

    1. **Total Hrs**: Actual hours worked from RotaCloud shifts (includes On-Call and Training hours)
    2. **Fixed Hrs**: Contracted weekly hours × (days in period ÷ 7) — **CALCULATED** (bold)
    3. **Hours**: MIN(Total - On-Call - Training, Fixed - On-Call - Training) — **CALCULATED** (bold)
    4. **Rate 1**: Employee's hourly rate from RotaCloud (or derived from annual salary)
    5. **Custom Role Columns**: Shifts with custom role rates appear as separate column pairs
    6. **On-Call Hrs**: Hours worked on On-Call shifts — from Attendance API (in_time/out_time) — **CALCULATED** (bold)
    7. **On-Call Shifts**: Count of assigned On-Call shifts (whether attended or not)
    8. **On-Call Flat Rate (£)**: £15.00 per assigned On-Call shift
    9. **Training Hrs**: Hours on Training-role shifts — from scheduled shift times — **CALCULATED** (bold)
    10. **Training Pay (£)**: Training Hrs × Rate 1 — **CALCULATED** (bold)
    11. **Overtime Hrs**: Total - Fixed — **CALCULATED** (bold)
    12. **Rate 2**: Overtime rate (default: UK minimum wage £12.21)
    13. **Holiday**: Approved holiday days and hours from RotaCloud

    ### Pay Calculation:
    - **Hours** (remaining fixed after on-call and training) at **Rate 1**
    - **Custom role hours** at their **custom rates**
    - **On-Call hours** at **Rate 1** + **£15 flat rate per shift**
    - **Training hours** at **Rate 1** (employee's own rate)
    - **Overtime** at **Rate 2**
    - **Holiday hours** at **Rate 1**

    **Formula**: `=(Hours×Rate1) + CustomRoles + (On-Call Hrs×Rate1) + (On-Call Shifts×£15) + (Training Hrs×Rate1) + (Overtime×Rate2) + (Holiday Hrs×Rate1)`

    ### Training vs On-Call:
    - **Training** hours use scheduled shift times (start/end) — no attendance lookup needed
    - **On-Call** hours use the Attendance API clock-in/out times
    - Both are excluded from the base "Hours" column to avoid double-counting
    - Training rows are highlighted in **light purple** (unless also On-Call, which takes priority)

    ### Row Highlighting:
    - 🟡 **Yellow** — staff with On-Call shifts
    - 🟣 **Purple** — staff with Training shifts only
    - 🟢 **Green** — Salaried staff
    - 🟠 **Orange** — Non-standard hourly rate
    """)
